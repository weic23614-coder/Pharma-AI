from __future__ import annotations

import argparse
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "app.db"


SKU_KEYS = {"sku", "sku_id", "商品编码", "产品编码", "商品sku", "商品id", "货号"}
NAME_KEYS = {"商品名称", "产品名称", "药品名称", "名称", "商品名", "通用名"}
CATEGORY_KEYS = {"类目", "商品类目", "一级类目", "二级类目", "品类", "科室"}
PRICE_KEYS = {"成交价", "单价", "实付单价", "吊牌价", "销售价", "销售单价", "gmv"}
COST_KEYS = {"成本", "采购价", "供货价", "成本价"}
QTY_KEYS = {"销量", "数量", "销售数量", "出库数量", "件数"}


def normalize_header(text: str) -> str:
    return str(text).strip().lower().replace(" ", "")


def find_col_idx(headers: list[str], candidates: set[str]) -> int | None:
    norm_candidates = {normalize_header(x) for x in candidates}
    for idx, h in enumerate(headers):
        if normalize_header(h) in norm_candidates:
            return idx
    return None


def to_float(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except Exception:
        return default


def load_rows(excel_path: Path, sheet_name: str | None = None) -> tuple[list[str], list[tuple]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    data = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
    return headers, data


def upsert_products(rows: list[dict]) -> int:
    now = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    count = 0
    for r in rows:
        cur.execute(
            """
            INSERT INTO products (
                sku_id, product_name, category, role, cost, original_price, gross_margin_rate, active, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(sku_id) DO UPDATE SET
                product_name=excluded.product_name,
                category=excluded.category,
                cost=excluded.cost,
                original_price=excluded.original_price,
                gross_margin_rate=excluded.gross_margin_rate,
                active=1,
                updated_at=excluded.updated_at
            """,
            (
                r["sku_id"],
                r["product_name"],
                r["category"],
                r["role"],
                r["cost"],
                r["original_price"],
                r["gross_margin_rate"],
                now,
            ),
        )
        count += 1
    conn.commit()
    conn.close()
    return count


def main() -> None:
    parser = argparse.ArgumentParser(description="Import yesterday sold products into products table.")
    parser.add_argument("--excel", required=True, help="Excel file path")
    parser.add_argument("--sheet", default=None, help="Sheet name (optional)")
    parser.add_argument("--default-role", default="main", choices=["main", "addon"], help="role if file has no role field")
    parser.add_argument("--default-cost-rate", type=float, default=0.78, help="fallback cost = price * rate when cost missing")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    headers, data = load_rows(excel_path, args.sheet)
    if not headers:
        raise SystemExit("Excel为空，无法导入。")

    sku_idx = find_col_idx(headers, SKU_KEYS)
    name_idx = find_col_idx(headers, NAME_KEYS)
    category_idx = find_col_idx(headers, CATEGORY_KEYS)
    price_idx = find_col_idx(headers, PRICE_KEYS)
    cost_idx = find_col_idx(headers, COST_KEYS)
    qty_idx = find_col_idx(headers, QTY_KEYS)

    missing = []
    if sku_idx is None:
        missing.append("SKU")
    if name_idx is None:
        missing.append("商品名称")
    if price_idx is None:
        missing.append("价格")
    if missing:
        print("当前表头:", headers)
        raise SystemExit(f"缺少关键字段: {', '.join(missing)}")

    merged: dict[str, dict] = {}
    for row in data:
        sku = str(row[sku_idx]).strip() if row[sku_idx] is not None else ""
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        if not sku or not name:
            continue
        category = str(row[category_idx]).strip() if category_idx is not None and row[category_idx] else "未分类"
        price = to_float(row[price_idx], 0)
        if price <= 0:
            continue
        cost = to_float(row[cost_idx], 0) if cost_idx is not None else 0
        if cost <= 0:
            cost = round(price * args.default_cost_rate, 2)
        qty = int(to_float(row[qty_idx], 1)) if qty_idx is not None else 1

        if sku not in merged:
            merged[sku] = {
                "sku_id": sku,
                "product_name": name,
                "category": category,
                "role": args.default_role,
                "cost": cost,
                "original_price": price,
                "qty": max(qty, 1),
            }
        else:
            merged[sku]["qty"] += max(qty, 1)
            merged[sku]["original_price"] = max(merged[sku]["original_price"], price)

    rows = []
    for v in merged.values():
        margin_rate = max(0.01, min(0.95, (v["original_price"] - v["cost"]) / v["original_price"]))
        rows.append(
            {
                "sku_id": v["sku_id"],
                "product_name": v["product_name"],
                "category": v["category"],
                "role": v["role"],
                "cost": round(v["cost"], 2),
                "original_price": round(v["original_price"], 2),
                "gross_margin_rate": round(margin_rate, 3),
            }
        )

    if not rows:
        raise SystemExit("没有可导入的商品行，请检查字段是否正确。")

    count = upsert_products(rows)
    print(f"导入完成: {count} 个SKU")
    print("示例:", rows[:5])


if __name__ == "__main__":
    main()
