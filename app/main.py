from __future__ import annotations

import asyncio
import io
import json
import os
import random
import sqlite3
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from starlette.requests import Request

from app.ai_brain import BailianAIBrain, _env_bool
from app.bundle_engine import BundleEngine, EngineCandidate, EngineInput, EnginePolicy, EngineStrategy

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "app.db"

app = FastAPI(title="1药网 AI 组货中间件 MVP", version="0.1.0")
templates = Jinja2Templates(directory=str(BASE_DIR / "app" / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "app" / "static")), name="static")
ai_brain = BailianAIBrain()


class MainItem(BaseModel):
    sku_id: str
    product_name: str
    category: str
    price: float
    cost: float


class CandidateItem(BaseModel):
    sku_id: str
    product_name: str
    cost: float
    original_price: float
    category: str | None = None


class RecommendRequest(BaseModel):
    user_intent: str = Field(default="checkout")
    main_item: MainItem
    candidate_pool: list[CandidateItem]
    user_id: str | None = None


def db_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS policies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            logic_type TEXT NOT NULL,
            prompt_hint TEXT NOT NULL,
            margin_rate REAL NOT NULL DEFAULT 0.35,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS cached_recommendations (
            cache_key TEXT PRIMARY KEY,
            payload TEXT NOT NULL,
            expires_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS recommendation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            main_sku_id TEXT NOT NULL,
            main_category TEXT NOT NULL,
            selected_sku_id TEXT,
            addon_price REAL,
            projected_profit REAL,
            source TEXT NOT NULL,
            latency_ms INTEGER NOT NULL,
            result_status TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            sku_id TEXT PRIMARY KEY,
            product_name TEXT NOT NULL,
            category TEXT NOT NULL,
            role TEXT NOT NULL,
            product_code TEXT,
            standard_code TEXT,
            manufacturer TEXT,
            department TEXT,
            cost REAL NOT NULL,
            original_price REAL NOT NULL,
            gross_margin_rate REAL NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS strategy_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            strategy_name TEXT NOT NULL,
            version TEXT NOT NULL,
            content_json TEXT NOT NULL,
            status TEXT NOT NULL,
            published_at TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS experiments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            exp_name TEXT NOT NULL,
            category TEXT NOT NULL,
            traffic_a REAL NOT NULL DEFAULT 0.5,
            traffic_b REAL NOT NULL DEFAULT 0.5,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS recommendation_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id TEXT NOT NULL,
            event_type TEXT NOT NULL,
            main_sku_id TEXT NOT NULL,
            selected_sku_id TEXT,
            variant TEXT,
            revenue REAL,
            margin REAL,
            created_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS upload_batches (
            id TEXT PRIMARY KEY,
            filename TEXT NOT NULL,
            total_rows INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS uploaded_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            sku_id TEXT NOT NULL,
            product_code TEXT,
            standard_code TEXT,
            product_name TEXT NOT NULL,
            manufacturer TEXT,
            category TEXT,
            business_mode TEXT,
            platform TEXT,
            merchant_name TEXT,
            price REAL NOT NULL,
            cost REAL NOT NULL,
            qty INTEGER NOT NULL DEFAULT 1,
            gmv REAL NOT NULL DEFAULT 0,
            revenue REAL NOT NULL DEFAULT 0,
            gaap_profit REAL NOT NULL DEFAULT 0,
            gaap_margin REAL NOT NULL DEFAULT 0,
            order_cnt INTEGER NOT NULL DEFAULT 0,
            customer_cnt INTEGER NOT NULL DEFAULT 0,
            arpo REAL NOT NULL DEFAULT 0,
            role_hint TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS bundle_recommendations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            main_sku_id TEXT NOT NULL,
            main_product_name TEXT NOT NULL,
            main_category TEXT NOT NULL,
            selected_sku_id TEXT NOT NULL,
            selected_product_name TEXT NOT NULL,
            medical_logic TEXT NOT NULL,
            addon_price REAL NOT NULL,
            projected_profit REAL NOT NULL,
            sales_copy TEXT NOT NULL,
            decision_payload TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS bundle_rules (
            main_sku_id TEXT PRIMARY KEY,
            main_product_name TEXT NOT NULL,
            selected_sku_id TEXT NOT NULL,
            selected_product_name TEXT NOT NULL,
            addon_price REAL NOT NULL,
            medical_logic TEXT NOT NULL,
            sales_copy TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pricing_recommendations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL,
            product_name TEXT NOT NULL,
            manufacturer TEXT,
            product_code TEXT,
            standard_code TEXT,
            business_mode TEXT,
            platform TEXT,
            merchant_name TEXT,
            current_price REAL NOT NULL,
            suggested_price REAL NOT NULL,
            delta_ratio REAL NOT NULL,
            sales_metric REAL NOT NULL,
            profit_metric REAL NOT NULL,
            current_margin REAL,
            predicted_margin REAL,
            confidence REAL NOT NULL DEFAULT 0.6,
            reason TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pricing_rules (
            rule_key TEXT PRIMARY KEY,
            product_name TEXT NOT NULL,
            manufacturer TEXT,
            product_code TEXT,
            standard_code TEXT,
            business_mode TEXT,
            platform TEXT,
            merchant_name TEXT,
            current_price REAL NOT NULL,
            suggested_price REAL NOT NULL,
            delta_ratio REAL NOT NULL,
            current_margin REAL,
            predicted_margin REAL,
            reason TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS llm_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            provider TEXT NOT NULL,
            model TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 1,
            monthly_budget_usd REAL NOT NULL DEFAULT 50,
            input_cost_per_1k REAL NOT NULL DEFAULT 0.001,
            output_cost_per_1k REAL NOT NULL DEFAULT 0.002,
            updated_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ai_usage_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scene TEXT NOT NULL,
            provider TEXT NOT NULL,
            model TEXT NOT NULL,
            prompt_tokens INTEGER NOT NULL DEFAULT 0,
            completion_tokens INTEGER NOT NULL DEFAULT 0,
            total_tokens INTEGER NOT NULL DEFAULT 0,
            estimated_cost_usd REAL NOT NULL DEFAULT 0,
            source TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    cols = [r["name"] for r in cur.execute("PRAGMA table_info(uploaded_products)").fetchall()]
    if "gmv" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN gmv REAL NOT NULL DEFAULT 0")
        conn.commit()
    if "product_code" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN product_code TEXT")
    if "standard_code" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN standard_code TEXT")
    if "manufacturer" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN manufacturer TEXT")
    if "business_mode" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN business_mode TEXT")
    if "platform" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN platform TEXT")
    if "merchant_name" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN merchant_name TEXT")
    if "revenue" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN revenue REAL NOT NULL DEFAULT 0")
    if "gaap_profit" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN gaap_profit REAL NOT NULL DEFAULT 0")
    if "gaap_margin" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN gaap_margin REAL NOT NULL DEFAULT 0")
    if "order_cnt" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN order_cnt INTEGER NOT NULL DEFAULT 0")
    if "customer_cnt" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN customer_cnt INTEGER NOT NULL DEFAULT 0")
    if "arpo" not in cols:
        cur.execute("ALTER TABLE uploaded_products ADD COLUMN arpo REAL NOT NULL DEFAULT 0")
    conn.commit()

    # Ensure products table has pricing/vender identifiers (for inventory upload + downstream matching).
    cols_p = [r["name"] for r in cur.execute("PRAGMA table_info(products)").fetchall()]
    if "product_code" not in cols_p:
        cur.execute("ALTER TABLE products ADD COLUMN product_code TEXT")
    if "standard_code" not in cols_p:
        cur.execute("ALTER TABLE products ADD COLUMN standard_code TEXT")
    if "manufacturer" not in cols_p:
        cur.execute("ALTER TABLE products ADD COLUMN manufacturer TEXT")
    if "department" not in cols_p:
        cur.execute("ALTER TABLE products ADD COLUMN department TEXT")
    conn.commit()

    cols_llm = [r["name"] for r in cur.execute("PRAGMA table_info(llm_settings)").fetchall()]
    if "api_key" not in cols_llm:
        cur.execute("ALTER TABLE llm_settings ADD COLUMN api_key TEXT")
    if "base_url" not in cols_llm:
        cur.execute("ALTER TABLE llm_settings ADD COLUMN base_url TEXT")
    if "timeout_sec" not in cols_llm:
        cur.execute("ALTER TABLE llm_settings ADD COLUMN timeout_sec REAL")
    conn.commit()

    cur.execute("SELECT COUNT(*) AS cnt FROM policies")
    count = cur.fetchone()["cnt"]
    if count == 0:
        now = datetime.now(timezone.utc).isoformat()
        seeds = [
            ("抗生素", "副作用对冲", "抗生素易破坏肠道菌群，优先推荐益生菌。", 0.35, 1, now),
            ("降脂药", "疗效协同", "他汀类可能消耗辅酶Q10，推荐辅酶Q10补充。", 0.38, 1, now),
            ("降糖药", "病因延展", "关注神经和血管并发风险，优先营养神经类。", 0.4, 1, now),
            ("高血压药", "慢病管理", "搭配血压计监测，强化管理闭环。", 0.42, 1, now),
        ]
        cur.executemany(
            """
            INSERT INTO policies (category, logic_type, prompt_hint, margin_rate, active, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            seeds,
        )
        conn.commit()
    cur.execute("SELECT COUNT(*) AS cnt FROM products")
    pcount = cur.fetchone()["cnt"]
    if pcount == 0:
        now = datetime.now(timezone.utc).isoformat()
        products = [
            ("A200", "盐酸二甲双胍片", "降糖药", "main", 23, 26, 0.115, 1, now),
            ("A300", "缬沙坦胶囊", "高血压药", "main", 27, 32, 0.156, 1, now),
            ("A123", "阿莫西林胶囊", "抗生素", "main", 22, 25, 0.12, 1, now),
            ("B801", "α-硫辛酸胶囊", "营养保健", "addon", 29, 99, 0.707, 1, now),
            ("B802", "家用血糖仪", "医疗器械", "addon", 65, 188, 0.654, 1, now),
            ("B901", "上臂式电子血压计", "医疗器械", "addon", 88, 259, 0.66, 1, now),
            ("B902", "辅酶Q10软胶囊", "营养保健", "addon", 36, 139, 0.741, 1, now),
            ("B001", "益生菌冻干粉", "营养保健", "addon", 30, 128, 0.766, 1, now),
        ]
        cur.executemany(
            """
            INSERT INTO products (
                sku_id, product_name, category, role, cost, original_price, gross_margin_rate, active, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            products,
        )
        conn.commit()
    cur.execute("SELECT COUNT(*) AS cnt FROM strategy_versions")
    scount = cur.fetchone()["cnt"]
    if scount == 0:
        now = datetime.now(timezone.utc).isoformat()
        strategy = {
            "title": "默认医嘱式推荐策略",
            "pricing_rules": {"anchor_ratio": 0.42, "min_margin_rate": 0.35},
            "copy_style": ["医学逻辑优先", "医嘱关怀语气", "不使用低价促销词"],
            "forbidden_terms": ["跳楼价", "秒杀", "白菜价", "清仓"],
        }
        cur.execute(
            """
            INSERT INTO strategy_versions (strategy_name, version, content_json, status, published_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ("smart-bundle-core", "v1.0.0", json.dumps(strategy, ensure_ascii=False), "published", now, now),
        )
        conn.commit()
    cur.execute("SELECT COUNT(*) AS cnt FROM experiments")
    ecount = cur.fetchone()["cnt"]
    if ecount == 0:
        now = datetime.now(timezone.utc).isoformat()
        cur.execute(
            """
            INSERT INTO experiments (exp_name, category, traffic_a, traffic_b, status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ("checkout-copy-pricing", "all", 0.5, 0.5, "running", now),
        )
        conn.commit()
    cur.execute("SELECT COUNT(*) AS cnt FROM llm_settings")
    lcount = cur.fetchone()["cnt"]
    if lcount == 0:
        cur.execute(
            """
            INSERT INTO llm_settings (
              provider, model, enabled, monthly_budget_usd, input_cost_per_1k, output_cost_per_1k, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            ("bailian", "qwen-plus", 1, 50, 0.0012, 0.0024, now_iso()),
        )
        conn.commit()
    conn.close()


@app.on_event("startup")
def startup_event() -> None:
    init_db()
    refresh_ai_brain_from_db()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def assign_variant(user_id: str | None) -> str:
    if not user_id:
        return "A" if random.random() < 0.5 else "B"
    return "A" if hash(user_id) % 2 == 0 else "B"


def get_latest_strategy() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    row = cur.execute(
        """
        SELECT * FROM strategy_versions
        WHERE status='published'
        ORDER BY published_at DESC, updated_at DESC
        LIMIT 1
        """
    ).fetchone()
    conn.close()
    if not row:
        return {"pricing_rules": {"anchor_ratio": 0.42, "min_margin_rate": 0.35}, "forbidden_terms": []}
    return json.loads(row["content_json"])


def select_candidates_by_pool_or_db(payload: RecommendRequest) -> list[CandidateItem]:
    if payload.candidate_pool:
        return payload.candidate_pool
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT sku_id, product_name, cost, original_price, category FROM products WHERE role='addon' AND active=1"
    ).fetchall()
    conn.close()
    return [CandidateItem(**dict(r)) for r in rows]


def infer_role(category: str | None, product_name: str) -> str:
    text = f"{category or ''}{product_name}"
    addon_tokens = ("保健", "器械", "血压计", "血糖仪", "辅酶", "益生菌", "维生素")
    return "addon" if any(t in text for t in addon_tokens) else "main"


def normalize_header(text: str) -> str:
    return str(text).strip().lower().replace(" ", "")


def find_col_idx(headers: list[str], candidates: set[str]) -> int | None:
    cand = {normalize_header(x) for x in candidates}
    for idx, h in enumerate(headers):
        if normalize_header(h) in cand:
            return idx
    return None


def to_float(v: Any, default: float = 0) -> float:
    if v in (None, ""):
        return default
    try:
        return float(v)
    except Exception:
        return default


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def _get_llm_setting() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    row = cur.execute("SELECT * FROM llm_settings ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return dict(row) if row else {}


def refresh_ai_brain_from_db() -> None:
    ai_brain.refresh_runtime(_get_llm_setting())


def _log_ai_usage(
    scene: str,
    model: str,
    usage: dict[str, int],
    source: str,
) -> None:
    setting = _get_llm_setting()
    in_cost = float(setting.get("input_cost_per_1k", 0.0012) or 0.0012)
    out_cost = float(setting.get("output_cost_per_1k", 0.0024) or 0.0024)
    p = int(usage.get("prompt_tokens", 0) or 0)
    c = int(usage.get("completion_tokens", 0) or 0)
    t = int(usage.get("total_tokens", p + c) or (p + c))
    est = round((p / 1000) * in_cost + (c / 1000) * out_cost, 6)
    try:
        conn = db_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO ai_usage_logs (
              scene, provider, model, prompt_tokens, completion_tokens, total_tokens, estimated_cost_usd, source, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (scene, "bailian", model, p, c, t, est, source, now_iso()),
        )
        conn.commit()
        conn.close()
    except Exception:
        # Usage logging should never block recommendation generation.
        return


def _log_ai_attempt(source: str, model: str) -> None:
    try:
        conn = db_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO ai_usage_logs (
              scene, provider, model, prompt_tokens, completion_tokens, total_tokens, estimated_cost_usd, source, created_at
            ) VALUES (?, ?, ?, 0, 0, 0, 0, ?, ?)
            """,
            ("bundle", "bailian", model, source, now_iso()),
        )
        conn.commit()
        conn.close()
    except Exception:
        return


def _is_ai_allowed() -> bool:
    setting = _get_llm_setting()
    if not bool(setting.get("enabled", 1)):
        return False
    key = (str(setting.get("api_key") or "").strip()) or os.getenv("BAILIAN_API_KEY", "").strip()
    if not key:
        return False
    enable_env = os.getenv("ENABLE_AI_BRAIN", None)
    if enable_env is not None and not _env_bool("ENABLE_AI_BRAIN", False):
        return False
    if not ai_brain.client:
        ai_brain.refresh_runtime(setting)
    return ai_brain.is_enabled() and bool(ai_brain.client)


def _build_recommendation_result(
    main_item: MainItem,
    user_id: str | None,
    candidates: list[CandidateItem],
    policy_like: dict[str, Any],
    strategy: dict[str, Any],
    variant: str,
    prefer_ai: bool = True,
    force_ai_only: bool = False,
    selection_counts: dict[str, int] | None = None,
    diversity_alpha: float = 0.003,
) -> dict[str, Any]:
    engine = BundleEngine()
    engine_input = EngineInput(
        main_sku_id=main_item.sku_id,
        main_product_name=main_item.product_name,
        main_category=main_item.category,
        main_price=main_item.price,
        main_cost=main_item.cost,
        user_id=user_id,
        variant=variant,
    )
    engine_policy = EnginePolicy(
        logic_type=policy_like["logic_type"],
        prompt_hint=policy_like["prompt_hint"],
        margin_rate=policy_like["margin_rate"],
    )
    engine_strategy = EngineStrategy(
        anchor_ratio=strategy.get("pricing_rules", {}).get("anchor_ratio", 0.42),
        min_margin_rate=strategy.get("pricing_rules", {}).get("min_margin_rate", 0.35),
        forbidden_terms=strategy.get("forbidden_terms", []),
    )
    engine_candidates = [
        EngineCandidate(
            sku_id=c.sku_id,
            product_name=c.product_name,
            cost=c.cost,
            original_price=c.original_price,
            category=c.category,
        )
        for c in candidates
    ]
    if prefer_ai and _is_ai_allowed():
        _log_ai_attempt("ai_attempt", ai_brain.model)
        ai_result = None
        try:
            ai_result = ai_brain.recommend(
            {
                "sku_id": main_item.sku_id,
                "product_name": main_item.product_name,
                "category": main_item.category,
                "price": main_item.price,
                "cost": main_item.cost,
            },
            {
                "logic_type": policy_like["logic_type"],
                "prompt_hint": policy_like["prompt_hint"],
                "margin_rate": policy_like["margin_rate"],
            },
            strategy,
            [
                {
                    "sku_id": c.sku_id,
                    "product_name": c.product_name,
                    "category": c.category,
                    "cost": c.cost,
                    "original_price": c.original_price,
                }
                for c in engine_candidates
            ],
            variant,
            )
        except Exception as e:
            _log_ai_attempt("ai_error", ai_brain.model)
            if force_ai_only:
                raise RuntimeError(f"AI调用异常: {str(e)}")
        if ai_result:
            llm_data = ai_result.get("result", {})
            usage = ai_result.get("usage", {})
            model_used = str(ai_result.get("model", ai_brain.model))
            _log_ai_usage("bundle", model_used, usage, "bailian_llm")

            # 批次生成启用多样性时：即使 LLM 给出了 selected_sku_id，
            # 也可能因为评分基准/候选分布导致“全都选同一个SKU”。
            # 为了保证你在后台看到的搭配不塌缩，这里直接走多样性规则选择。
            if selection_counts is not None:
                return engine.recommend(
                    engine_input=engine_input,
                    policy=engine_policy,
                    strategy=engine_strategy,
                    candidates=engine_candidates,
                    selection_counts=selection_counts,
                    diversity_alpha=diversity_alpha,
                )

            candidate_map = {c.sku_id: c for c in engine_candidates}
            selected = candidate_map.get(str(llm_data.get("selected_sku_id", "")).strip())
            if selected:
                anchor_ratio = strategy.get("pricing_rules", {}).get("anchor_ratio", 0.42)
                min_margin_rate = strategy.get("pricing_rules", {}).get("min_margin_rate", 0.35)
                margin_rate = max(float(policy_like["margin_rate"]), float(min_margin_rate))
                ar = anchor_ratio if variant == "A" else max(0.3, anchor_ratio - 0.03)
                addon_price = round(max(selected.cost * (1 + margin_rate), selected.original_price * ar), 2)
                projected_profit = round(addon_price - selected.cost, 2)
                forbidden_terms = strategy.get("forbidden_terms", [])
                sales_copy = str(llm_data.get("sales_copy") or "").strip()
                if not sales_copy:
                    sales_copy = engine.combo_sales_copy(
                        main_item.product_name,
                        main_item.category,
                        selected,
                        engine_policy,
                        variant,
                        forbidden_terms,
                        main_sku_id=main_item.sku_id,
                    )
                for bad in forbidden_terms:
                    sales_copy = sales_copy.replace(str(bad), "")
                return {
                    "recommendation": {
                        "request_id": f"req_{int(datetime.now().timestamp() * 1000)}_{random.randint(100, 999)}",
                        "variant": variant,
                        "selected_sku_id": selected.sku_id,
                        "product_name": selected.product_name,
                        "medical_logic": str(llm_data.get("medical_logic") or policy_like["logic_type"]),
                        "sales_copy": sales_copy,
                        "pricing_strategy": {
                            "addon_price": addon_price,
                            "original_price": selected.original_price,
                            "display_tag": f"加{addon_price:.0f}元换购价",
                        },
                        "projected_profit": projected_profit,
                        "decision_trace": {
                            "source": "bailian_llm",
                            "confidence": float(llm_data.get("confidence", 0.5) or 0.5),
                            "medical_reason": str(llm_data.get("medical_reason", "")),
                            "model": model_used,
                        },
                    }
                }
        if force_ai_only:
            raise RuntimeError("AI未返回可用结果（可能超时或返回SKU不在候选池）。")
    elif force_ai_only:
        raise RuntimeError("AI当前未启用。")
    return engine.recommend(
        engine_input=engine_input,
        policy=engine_policy,
        strategy=engine_strategy,
        candidates=engine_candidates,
        selection_counts=selection_counts,
        diversity_alpha=diversity_alpha,
    )


async def run_ai_logic(payload: RecommendRequest, variant: str) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM policies WHERE category=? AND active=1 ORDER BY updated_at DESC LIMIT 1",
        (payload.main_item.category,),
    )
    policy = cur.fetchone()
    conn.close()
    if not policy:
        raise HTTPException(status_code=404, detail="当前类目没有启用策略")
    candidates = select_candidates_by_pool_or_db(payload)
    if not candidates:
        raise HTTPException(status_code=400, detail="候选池为空")
    strategy = get_latest_strategy()
    try:
        return await asyncio.to_thread(
            _build_recommendation_result,
            payload.main_item,
            payload.user_id,
            candidates,
            {"logic_type": policy["logic_type"], "prompt_hint": policy["prompt_hint"], "margin_rate": policy["margin_rate"]},
            strategy,
            variant,
            True,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def write_log(
    req: RecommendRequest,
    result: dict[str, Any] | None,
    source: str,
    latency_ms: int,
    result_status: str,
) -> None:
    conn = db_conn()
    cur = conn.cursor()
    rec = result.get("recommendation", {}) if result else {}
    cur.execute(
        """
        INSERT INTO recommendation_logs (
            main_sku_id, main_category, selected_sku_id, addon_price,
            projected_profit, source, latency_ms, result_status, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            req.main_item.sku_id,
            req.main_item.category,
            rec.get("selected_sku_id"),
            rec.get("pricing_strategy", {}).get("addon_price"),
            rec.get("projected_profit"),
            source,
            latency_ms,
            result_status,
            now_iso(),
        ),
    )
    conn.commit()
    conn.close()


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/admin/ai-status")
def ai_status() -> dict[str, Any]:
    setting = _get_llm_setting()
    return {
        "enabled": _is_ai_allowed(),
        "model": setting.get("model", ai_brain.model),
        "base_url": ai_brain.base_url,
        "timeout_sec": ai_brain.timeout,
    }


@app.get("/api/admin/db-info")
def db_info() -> dict[str, Any]:
    # Debug helper: confirm which app.db instance the running server is using.
    info: dict[str, Any] = {
        "db_path": str(DB_PATH),
        "db_exists": DB_PATH.exists(),
    }
    try:
        if DB_PATH.exists():
            st = DB_PATH.stat()
            info["db_mtime"] = st.st_mtime
    except Exception:
        pass
    try:
        conn = db_conn()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT batch_id, MAX(updated_at) AS max_updated FROM bundle_recommendations GROUP BY batch_id ORDER BY max_updated DESC LIMIT 1"
        ).fetchone()
        info["latest_bundle_by_updated_at"] = {"batch_id": row[0], "max_updated": row[1]} if row else None
        info["bundle_recommendations_total"] = cur.execute("SELECT COUNT(*) FROM bundle_recommendations").fetchone()[0]
        info["upload_batches_total"] = cur.execute("SELECT COUNT(*) FROM upload_batches").fetchone()[0]
        conn.close()
    except Exception:
        # Debug endpoint should never crash the server.
        info["db_query_error"] = True
    return info


@app.get("/api/admin/ai-ping")
def ai_ping() -> dict[str, Any]:
    if not _is_ai_allowed():
        return {"ok": False, "reason": "ai_disabled"}
    try:
        # Minimal model reachability probe.
        res = ai_brain.recommend(
            {"sku_id": "PING001", "product_name": "测试商品A", "category": "测试", "price": 10, "cost": 8},
            {"logic_type": "测试逻辑", "prompt_hint": "测试提示", "margin_rate": 0.35},
            {"pricing_rules": {"anchor_ratio": 0.42, "min_margin_rate": 0.35}, "forbidden_terms": []},
            [{"sku_id": "PING002", "product_name": "测试商品B", "category": "测试", "cost": 5, "original_price": 20}],
            "A",
        )
        if not res:
            return {"ok": False, "reason": "empty_response"}
        return {"ok": True, "model": ai_brain.model}
    except Exception as exc:
        return {"ok": False, "reason": str(exc)}


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "admin.html", {})


@app.get("/api/admin/policies")
def list_policies() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute("SELECT * FROM policies ORDER BY id DESC").fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}


class PolicyIn(BaseModel):
    category: str
    logic_type: str
    prompt_hint: str
    margin_rate: float
    active: bool = True


@app.post("/api/admin/policies")
def create_policy(data: PolicyIn) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO policies (category, logic_type, prompt_hint, margin_rate, active, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            data.category,
            data.logic_type,
            data.prompt_hint,
            data.margin_rate,
            1 if data.active else 0,
            now_iso(),
        ),
    )
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return {"id": new_id, "message": "created"}


@app.put("/api/admin/policies/{policy_id}")
def update_policy(policy_id: int, data: PolicyIn) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE policies
        SET category=?, logic_type=?, prompt_hint=?, margin_rate=?, active=?, updated_at=?
        WHERE id=?
        """,
        (
            data.category,
            data.logic_type,
            data.prompt_hint,
            data.margin_rate,
            1 if data.active else 0,
            now_iso(),
            policy_id,
        ),
    )
    conn.commit()
    conn.close()
    return {"message": "updated"}


@app.delete("/api/admin/policies/{policy_id}")
def delete_policy(policy_id: int) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM policies WHERE id=?", (policy_id,))
    conn.commit()
    conn.close()
    return {"message": "deleted"}


@app.get("/api/admin/metrics")
def metrics() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) AS cnt FROM recommendation_logs").fetchone()["cnt"]
    success = cur.execute(
        "SELECT COUNT(*) AS cnt FROM recommendation_logs WHERE result_status='ok'"
    ).fetchone()["cnt"]
    cache_hits = cur.execute(
        "SELECT COUNT(*) AS cnt FROM recommendation_logs WHERE source='cache'"
    ).fetchone()["cnt"]
    avg_profit = cur.execute(
        "SELECT COALESCE(AVG(projected_profit), 0) AS val FROM recommendation_logs"
    ).fetchone()["val"]
    ctr = cur.execute(
        """
        SELECT
        CAST(SUM(CASE WHEN event_type='click' THEN 1 ELSE 0 END) AS REAL) /
        NULLIF(SUM(CASE WHEN event_type='exposure' THEN 1 ELSE 0 END), 0) AS v
        FROM recommendation_events
        """
    ).fetchone()["v"]
    cvr = cur.execute(
        """
        SELECT
        CAST(SUM(CASE WHEN event_type='order_addon' THEN 1 ELSE 0 END) AS REAL) /
        NULLIF(SUM(CASE WHEN event_type='click' THEN 1 ELSE 0 END), 0) AS v
        FROM recommendation_events
        """
    ).fetchone()["v"]
    revenue = cur.execute(
        "SELECT COALESCE(SUM(revenue), 0) AS v FROM recommendation_events WHERE event_type='order_addon'"
    ).fetchone()["v"]
    margin = cur.execute(
        "SELECT COALESCE(SUM(margin), 0) AS v FROM recommendation_events WHERE event_type='order_addon'"
    ).fetchone()["v"]
    ai_total = cur.execute(
        "SELECT COUNT(*) AS cnt FROM recommendation_logs WHERE source LIKE 'ai_%'"
    ).fetchone()["cnt"]
    ai_success = cur.execute(
        "SELECT COUNT(*) AS cnt FROM recommendation_logs WHERE source='ai_realtime' AND result_status='ok'"
    ).fetchone()["cnt"]
    ai_fallback = cur.execute(
        "SELECT COUNT(*) AS cnt FROM recommendation_logs WHERE source='fallback' AND result_status='timeout_or_error'"
    ).fetchone()["cnt"]
    token = cur.execute("SELECT COALESCE(SUM(total_tokens),0) AS v FROM ai_usage_logs").fetchone()["v"]
    cost = cur.execute("SELECT COALESCE(SUM(estimated_cost_usd),0) AS v FROM ai_usage_logs").fetchone()["v"]
    conn.close()
    hit_rate = round(cache_hits / total, 3) if total else 0
    success_rate = round(success / total, 3) if total else 0
    return {
        "total_requests": total,
        "success_rate": success_rate,
        "cache_hit_rate": hit_rate,
        "avg_projected_profit": round(avg_profit, 2),
        "ctr": round(ctr or 0, 3),
        "cvr": round(cvr or 0, 3),
        "addon_revenue": round(revenue or 0, 2),
        "addon_margin": round(margin or 0, 2),
        "ai_total_calls": ai_total,
        "ai_success_rate": round((ai_success / ai_total), 3) if ai_total else 0,
        "ai_fallback_rate": round((ai_fallback / max(total, 1)), 3),
        "ai_total_tokens": int(token or 0),
        "ai_estimated_cost_usd": round(float(cost or 0), 6),
    }


@app.post("/api/recommend")
async def recommend(req: RecommendRequest) -> dict[str, Any]:
    started = datetime.now()
    cache_key = f"{req.main_item.sku_id}:{req.main_item.category}"
    conn = db_conn()
    cur = conn.cursor()
    row = cur.execute(
        "SELECT payload, expires_at FROM cached_recommendations WHERE cache_key=?",
        (cache_key,),
    ).fetchone()
    conn.close()

    if row and datetime.fromisoformat(row["expires_at"]) > datetime.now(timezone.utc):
        result = json.loads(row["payload"])
        latency = int((datetime.now() - started).total_seconds() * 1000)
        write_log(req, result, "cache", latency, "ok")
        return result

    variant = assign_variant(req.user_id)
    try:
        result = await asyncio.wait_for(run_ai_logic(req, variant), timeout=1.5)
        latency = int((datetime.now() - started).total_seconds() * 1000)
        trace_source = (
            result.get("recommendation", {})
            .get("decision_trace", {})
            .get("source", "rule_engine")
        )
        log_source = "ai_realtime" if trace_source == "bailian_llm" else "rule_realtime"
        write_log(req, result, log_source, latency, "ok")
    except Exception:
        latency = int((datetime.now() - started).total_seconds() * 1000)
        write_log(req, None, "fallback", latency, "timeout_or_error")
        return {"recommendation": None, "fallback": "skip_module"}

    if req.main_item.category in {"抗生素", "降糖药", "高血压药", "降脂药"}:
        conn = db_conn()
        cur = conn.cursor()
        expires = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        cur.execute(
            """
            INSERT INTO cached_recommendations (cache_key, payload, expires_at)
            VALUES (?, ?, ?)
            ON CONFLICT(cache_key) DO UPDATE SET payload=excluded.payload, expires_at=excluded.expires_at
            """,
            (cache_key, json.dumps(result, ensure_ascii=False), expires),
        )
        conn.commit()
        conn.close()

    return result


class EventIn(BaseModel):
    request_id: str
    event_type: str
    main_sku_id: str
    selected_sku_id: str | None = None
    variant: str | None = None
    revenue: float | None = None
    margin: float | None = None


@app.post("/api/events")
def ingest_event(event: EventIn) -> dict[str, str]:
    if event.event_type not in {"exposure", "click", "order_addon"}:
        raise HTTPException(status_code=400, detail="invalid event_type")
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO recommendation_events (
            request_id, event_type, main_sku_id, selected_sku_id, variant, revenue, margin, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            event.request_id,
            event.event_type,
            event.main_sku_id,
            event.selected_sku_id,
            event.variant,
            event.revenue,
            event.margin,
            now_iso(),
        ),
    )
    conn.commit()
    conn.close()
    return {"message": "ok"}


@app.get("/api/admin/products")
def list_products(role: str | None = Query(default=None)) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    if role:
        rows = cur.execute("SELECT * FROM products WHERE role=? ORDER BY updated_at DESC", (role,)).fetchall()
    else:
        rows = cur.execute("SELECT * FROM products ORDER BY updated_at DESC").fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}


class ProductIn(BaseModel):
    sku_id: str
    product_name: str
    category: str
    role: str
    cost: float
    original_price: float
    gross_margin_rate: float
    active: bool = True


@app.post("/api/admin/products")
def upsert_product(data: ProductIn) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO products (
            sku_id, product_name, category, role, cost, original_price, gross_margin_rate, active, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(sku_id) DO UPDATE SET
            product_name=excluded.product_name,
            category=excluded.category,
            role=excluded.role,
            cost=excluded.cost,
            original_price=excluded.original_price,
            gross_margin_rate=excluded.gross_margin_rate,
            active=excluded.active,
            updated_at=excluded.updated_at
        """,
        (
            data.sku_id,
            data.product_name,
            data.category,
            data.role,
            data.cost,
            data.original_price,
            data.gross_margin_rate,
            1 if data.active else 0,
            now_iso(),
        ),
    )
    conn.commit()
    conn.close()
    return {"message": "saved"}


@app.get("/api/admin/strategies")
def list_strategies() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute("SELECT * FROM strategy_versions ORDER BY updated_at DESC").fetchall()
    conn.close()
    items = []
    for r in rows:
        d = dict(r)
        d["content_json"] = json.loads(d["content_json"])
        items.append(d)
    return {"items": items}


class StrategyIn(BaseModel):
    strategy_name: str
    version: str
    content_json: dict[str, Any]
    status: str = "draft"


@app.post("/api/admin/strategies")
def create_strategy(data: StrategyIn) -> dict[str, Any]:
    if data.status not in {"draft", "published"}:
        raise HTTPException(status_code=400, detail="invalid status")
    conn = db_conn()
    cur = conn.cursor()
    now = now_iso()
    pub = now if data.status == "published" else None
    cur.execute(
        """
        INSERT INTO strategy_versions (strategy_name, version, content_json, status, published_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (data.strategy_name, data.version, json.dumps(data.content_json, ensure_ascii=False), data.status, pub, now),
    )
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return {"id": rid, "message": "created"}


@app.post("/api/admin/strategies/{strategy_id}/publish")
def publish_strategy(strategy_id: int) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    row = cur.execute("SELECT strategy_name FROM strategy_versions WHERE id=?", (strategy_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="strategy not found")
    strategy_name = row["strategy_name"]
    now = now_iso()
    cur.execute(
        "UPDATE strategy_versions SET status='draft' WHERE strategy_name=? AND status='published'",
        (strategy_name,),
    )
    cur.execute(
        "UPDATE strategy_versions SET status='published', published_at=?, updated_at=? WHERE id=?",
        (now, now, strategy_id),
    )
    conn.commit()
    conn.close()
    return {"message": "published"}


@app.get("/api/admin/experiments")
def list_experiments() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute("SELECT * FROM experiments ORDER BY updated_at DESC").fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}


class ExperimentIn(BaseModel):
    exp_name: str
    category: str = "all"
    traffic_a: float = 0.5
    traffic_b: float = 0.5
    status: str = "running"


@app.post("/api/admin/experiments")
def upsert_experiment(data: ExperimentIn) -> dict[str, str]:
    if round(data.traffic_a + data.traffic_b, 5) != 1:
        raise HTTPException(status_code=400, detail="traffic_a + traffic_b must equal 1")
    if data.status not in {"running", "paused"}:
        raise HTTPException(status_code=400, detail="invalid status")
    conn = db_conn()
    cur = conn.cursor()
    row = cur.execute("SELECT id FROM experiments WHERE exp_name=?", (data.exp_name,)).fetchone()
    if row:
        cur.execute(
            """
            UPDATE experiments
            SET category=?, traffic_a=?, traffic_b=?, status=?, updated_at=?
            WHERE exp_name=?
            """,
            (data.category, data.traffic_a, data.traffic_b, data.status, now_iso(), data.exp_name),
        )
    else:
        cur.execute(
            """
            INSERT INTO experiments (exp_name, category, traffic_a, traffic_b, status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (data.exp_name, data.category, data.traffic_a, data.traffic_b, data.status, now_iso()),
        )
    conn.commit()
    conn.close()
    return {"message": "saved"}


@app.get("/api/admin/ab-report")
def ab_report() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT
          variant,
          SUM(CASE WHEN event_type='exposure' THEN 1 ELSE 0 END) AS exposure,
          SUM(CASE WHEN event_type='click' THEN 1 ELSE 0 END) AS click,
          SUM(CASE WHEN event_type='order_addon' THEN 1 ELSE 0 END) AS orders,
          COALESCE(SUM(CASE WHEN event_type='order_addon' THEN margin ELSE 0 END), 0) AS margin
        FROM recommendation_events
        WHERE variant IS NOT NULL
        GROUP BY variant
        """
    ).fetchall()
    conn.close()
    items = []
    for r in rows:
        d = dict(r)
        exp = d["exposure"] or 0
        clk = d["click"] or 0
        d["ctr"] = round((clk / exp), 3) if exp else 0
        d["cvr"] = round((d["orders"] / clk), 3) if clk else 0
        items.append(d)
    return {"items": items}


@app.post("/api/ops/upload-catalog")
async def ops_upload_catalog(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件(.xlsx)")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖") from exc

    content = await file.read()
    # 不用 read_only：部分 BI/药网导出表在 read_only 下会错误只解析出首列，导致缺列误报
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件为空")

    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    sku_idx = find_col_idx(headers, {"sku", "sku_id", "商品sku", "商品id", "货号"})
    product_code_idx = find_col_idx(headers, {"商品编码", "产品编码", "外部编码", "编码"})
    standard_code_idx = find_col_idx(headers, {"标品主码", "标准品主码", "标准码", "spu", "spu编码", "标品编码"})
    name_idx = find_col_idx(headers, {"商品名称", "产品名称", "药品名称", "名称", "商品名", "通用名"})
    mfr_idx = find_col_idx(headers, {"生产厂商", "厂家", "生产厂家", "厂商"})
    cat_idx = find_col_idx(headers, {"类目", "商品类目", "一级类目", "二级类目", "品类", "科室"})
    biz_mode_idx = find_col_idx(headers, {"业务模式", "模式"})
    platform_idx = find_col_idx(headers, {"平台", "渠道", "销售渠道"})
    merchant_idx = find_col_idx(headers, {"商家名称", "店铺名称", "商家", "店铺"})
    price_idx = find_col_idx(headers, {"成交价", "单价", "实付单价", "吊牌价", "销售价", "销售单价", "零售价", "指导价", "gmv"})
    cost_idx = find_col_idx(headers, {"成本", "采购价", "供货价", "成本价"})
    qty_idx = find_col_idx(headers, {"销量", "数量", "销售数量", "出库数量", "件数"})
    gmv_idx = find_col_idx(headers, {"gmv", "销售额", "订单金额"})
    revenue_idx = find_col_idx(headers, {"revenue", "rev", "营收", "收入"})
    gaap_profit_idx = find_col_idx(headers, {"gaap毛利额(去税)", "gaap毛利额", "毛利额", "毛利"})
    gaap_margin_idx = find_col_idx(headers, {"gaap毛利率(去税)(%)", "gaap毛利率", "毛利率"})
    order_cnt_idx = find_col_idx(headers, {"订单数", "订单量"})
    customer_cnt_idx = find_col_idx(headers, {"顾客数", "用户数", "买家数"})
    arpo_idx = find_col_idx(headers, {"arpo", "客单价"})

    missing = []
    # 允许：仅标品主码（standard_code）也可作为“识别锚点”
    if sku_idx is None and product_code_idx is None and standard_code_idx is None:
        missing.append("SKU")
    if name_idx is None:
        missing.append("商品名称")
    # 运营约定：待组货清单可不提供「价格」，只关心 SKU 碰撞；缺失时用成本反推或占位价入库（仍满足 DB 非空约束）。
    if missing:
        head_preview = "、".join(h for h in headers[:30] if h) or "(空)"
        raise HTTPException(
            status_code=400,
            detail=f"缺少关键字段: {', '.join(missing)}。当前表头前若干列：{head_preview}",
        )

    batch_id = f"batch_{uuid.uuid4().hex[:10]}"
    now = now_iso()
    parsed = []
    for row in rows[1:]:
        sku = str(row[sku_idx]).strip() if sku_idx is not None and row[sku_idx] is not None else ""
        product_code = str(row[product_code_idx]).strip() if product_code_idx is not None and row[product_code_idx] is not None else ""
        if not sku:
            sku = product_code
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        if not sku or not name:
            continue
        standard_code = str(row[standard_code_idx]).strip() if standard_code_idx is not None and row[standard_code_idx] else ""
        if not sku:
            sku = standard_code
        manufacturer = str(row[mfr_idx]).strip() if mfr_idx is not None and row[mfr_idx] else ""
        category = str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] else "未分类"
        business_mode = str(row[biz_mode_idx]).strip() if biz_mode_idx is not None and row[biz_mode_idx] else ""
        platform = str(row[platform_idx]).strip() if platform_idx is not None and row[platform_idx] else ""
        merchant_name = str(row[merchant_idx]).strip() if merchant_idx is not None and row[merchant_idx] else ""
        price = to_float(row[price_idx], 0) if price_idx is not None else 0.0
        cost = to_float(row[cost_idx], 0) if cost_idx is not None else 0.0
        if price <= 0 and cost > 0:
            price = round(cost / 0.78, 2)
        elif price <= 0 and cost <= 0:
            price = 1.0
            cost = round(price * 0.78, 2)
        elif cost <= 0:
            cost = round(price * 0.78, 2)
        qty = int(to_float(row[qty_idx], 1)) if qty_idx is not None else 1
        gmv = to_float(row[gmv_idx], 0) if gmv_idx is not None else round(price * max(qty, 1), 2)
        revenue = to_float(row[revenue_idx], 0) if revenue_idx is not None else 0.0
        gaap_profit = to_float(row[gaap_profit_idx], 0) if gaap_profit_idx is not None else 0.0
        gaap_margin = to_float(row[gaap_margin_idx], 0) if gaap_margin_idx is not None else 0.0
        order_cnt = int(to_float(row[order_cnt_idx], qty)) if order_cnt_idx is not None else int(max(qty, 1))
        customer_cnt = int(to_float(row[customer_cnt_idx], 0)) if customer_cnt_idx is not None else 0
        arpo = to_float(row[arpo_idx], 0) if arpo_idx is not None else 0.0
        parsed.append(
            (
                batch_id,
                sku,
                product_code,
                standard_code,
                name,
                manufacturer,
                category,
                business_mode,
                platform,
                merchant_name,
                price,
                cost,
                max(qty, 1),
                gmv,
                revenue,
                gaap_profit,
                gaap_margin,
                max(order_cnt, 0),
                max(customer_cnt, 0),
                arpo,
                infer_role(category, name),
                now,
            )
        )

    if not parsed:
        raise HTTPException(status_code=400, detail="没有可导入的数据行")

    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO upload_batches (id, filename, total_rows, created_at) VALUES (?, ?, ?, ?)",
        (batch_id, file.filename, len(parsed), now),
    )
    cur.executemany(
        """
        INSERT INTO uploaded_products (
          batch_id, sku_id, product_code, standard_code, product_name, manufacturer, category, business_mode,
          platform, merchant_name, price, cost, qty, gmv, revenue, gaap_profit, gaap_margin,
          order_cnt, customer_cnt, arpo, role_hint, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        parsed,
    )
    conn.commit()
    conn.close()
    return {"batch_id": batch_id, "total_rows": len(parsed)}


@app.post("/api/ops/library/upload")
async def ops_upload_library(
    file: UploadFile = File(...),
    default_role: str = Query(default="addon"),
) -> dict[str, Any]:
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="仅支持 Excel 文件(.xlsx)")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖") from exc
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件为空")
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    sku_idx = find_col_idx(headers, {"sku", "sku_id", "商品编码", "产品编码", "商品sku", "商品id", "货号"})
    product_code_idx = find_col_idx(headers, {"商品编码", "产品编码", "外部编码", "编码"})
    standard_code_idx = find_col_idx(headers, {"标品主码", "标准品主码", "标准码", "spu", "spu编码", "标品编码"})
    name_idx = find_col_idx(headers, {"商品名称", "产品名称", "药品名称", "名称", "商品名", "通用名"})
    cat_idx = find_col_idx(headers, {"类目", "商品类目", "一级类目", "二级类目", "品类", "科室"})
    mfr_idx = find_col_idx(headers, {"生产厂商", "厂家", "生产厂家", "厂商"})
    price_idx = find_col_idx(headers, {"成交价", "单价", "实付单价", "吊牌价", "销售价", "销售单价", "gmv"})
    cost_idx = find_col_idx(headers, {"成本", "采购价", "供货价", "成本价", "revenue"})
    role_idx = find_col_idx(headers, {"角色", "role", "商品角色", "类型"})
    missing = []
    if sku_idx is None and product_code_idx is None and standard_code_idx is None:
        missing.append("SKU(或商品编码/标品主码)")
    if name_idx is None:
        missing.append("商品名称")
    if price_idx is None:
        missing.append("价格")
    if missing:
        head_preview = "、".join(headers[:30]) or "(空)"
        raise HTTPException(
            status_code=400,
            detail=f"缺少关键字段: {', '.join(missing)}。当前表头：{head_preview}",
        )
    now = now_iso()
    upserts = []
    for row in rows[1:]:
        sku = ""
        if sku_idx is not None and row[sku_idx] is not None:
            sku = str(row[sku_idx]).strip()
        product_code = str(row[product_code_idx]).strip() if product_code_idx is not None and row[product_code_idx] is not None else ""
        standard_code = str(row[standard_code_idx]).strip() if standard_code_idx is not None and row[standard_code_idx] is not None else ""
        if not sku:
            sku = product_code or standard_code
        if not sku:
            continue
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        if not sku or not name:
            continue
        category = str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] else "未分类"
        manufacturer = str(row[mfr_idx]).strip() if mfr_idx is not None and row[mfr_idx] is not None else ""
        price = to_float(row[price_idx], 0)
        if price <= 0:
            continue
        cost = to_float(row[cost_idx], 0) if cost_idx is not None else 0
        if cost <= 0:
            cost = round(price * 0.78, 2)
        role_val = str(row[role_idx]).strip().lower() if role_idx is not None and row[role_idx] else ""
        role = role_val if role_val in {"main", "addon"} else default_role
        if role not in {"main", "addon"}:
            role = infer_role(category, name)
        margin_rate = max(0.01, min(0.95, (price - cost) / max(price, 1)))
        # 若未提供“产品编码”，用“标品主码”兜底填到产品编码列（供页面展示与运营检索）
        if not product_code and standard_code:
            product_code = standard_code
        upserts.append(
            (sku, name, category, role, product_code, standard_code, manufacturer, cost, price, round(margin_rate, 3), 1, now)
        )
    if not upserts:
        raise HTTPException(status_code=400, detail="没有可导入商品")
    conn = db_conn()
    cur = conn.cursor()
    cur.executemany(
        """
        INSERT INTO products (
          sku_id, product_name, category, role, product_code, standard_code, manufacturer,
          cost, original_price, gross_margin_rate, active, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(sku_id) DO UPDATE SET
          product_name=excluded.product_name,
          category=excluded.category,
          role=excluded.role,
          product_code=excluded.product_code,
          standard_code=excluded.standard_code,
          manufacturer=excluded.manufacturer,
          cost=excluded.cost,
          original_price=excluded.original_price,
          gross_margin_rate=excluded.gross_margin_rate,
          active=excluded.active,
          updated_at=excluded.updated_at
        """,
        upserts,
    )
    conn.commit()
    main_cnt = cur.execute("SELECT COUNT(*) AS c FROM products WHERE role='main' AND active=1").fetchone()["c"]
    addon_cnt = cur.execute("SELECT COUNT(*) AS c FROM products WHERE role='addon' AND active=1").fetchone()["c"]
    conn.close()
    return {"imported": len(upserts), "main_count": main_cnt, "addon_count": addon_cnt}


@app.get("/api/ops/library/stats")
def ops_library_stats() -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) AS c FROM products WHERE active=1").fetchone()["c"]
    main_cnt = cur.execute("SELECT COUNT(*) AS c FROM products WHERE role='main' AND active=1").fetchone()["c"]
    addon_cnt = cur.execute("SELECT COUNT(*) AS c FROM products WHERE role='addon' AND active=1").fetchone()["c"]
    conn.close()
    return {"total": total, "main_count": main_cnt, "addon_count": addon_cnt}


@app.get("/api/ops/inventory/list")
def ops_inventory_list(
    q: str | None = Query(default=None),
    role: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    where = ["active=1"]
    params: list[Any] = []
    if q:
        where.append("(sku_id LIKE ? OR product_name LIKE ? OR category LIKE ?)")
        like = f"%{q}%"
        params.extend([like, like, like])
    if role in {"main", "addon"}:
        where.append("role=?")
        params.append(role)
    where_sql = " AND ".join(where)
    total = cur.execute(f"SELECT COUNT(*) AS c FROM products WHERE {where_sql}", tuple(params)).fetchone()["c"]
    rows = cur.execute(
        f"""
        SELECT
          sku_id, product_name, category, role, product_code, standard_code, manufacturer,
          category AS department,
          cost, original_price, gross_margin_rate, active, updated_at
        FROM products
        WHERE {where_sql}
        ORDER BY updated_at DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [limit, offset]),
    ).fetchall()
    conn.close()
    return {"total": total, "items": [dict(r) for r in rows]}


@app.post("/api/ops/inventory/upload")
async def ops_inventory_upload(
    file: UploadFile = File(...),
    default_role: str = Query(default="addon"),
) -> dict[str, Any]:
    return await ops_upload_library(file=file, default_role=default_role)


@app.get("/api/admin/budget")
def admin_budget() -> dict[str, Any]:
    setting = _get_llm_setting()
    conn = db_conn()
    cur = conn.cursor()
    usage = cur.execute(
        """
        SELECT
          COALESCE(SUM(prompt_tokens),0) AS prompt_tokens,
          COALESCE(SUM(completion_tokens),0) AS completion_tokens,
          COALESCE(SUM(total_tokens),0) AS total_tokens,
          COALESCE(SUM(estimated_cost_usd),0) AS estimated_cost
        FROM ai_usage_logs
        """
    ).fetchone()
    conn.close()
    monthly_budget = float(setting.get("monthly_budget_usd", 50) or 50)
    estimated = float(usage["estimated_cost"] or 0)
    rate = round((estimated / monthly_budget), 4) if monthly_budget > 0 else 0
    dash = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    db_base = (str(setting.get("base_url") or "").strip()) or dash
    db_timeout = setting.get("timeout_sec")
    try:
        timeout_out = float(db_timeout) if db_timeout is not None else float(ai_brain.timeout)
    except (TypeError, ValueError):
        timeout_out = float(ai_brain.timeout)
    key_cfg = bool(
        (str(setting.get("api_key") or "").strip()) or os.getenv("BAILIAN_API_KEY", "").strip()
    )
    return {
        "provider": setting.get("provider", "bailian"),
        "model": setting.get("model", ai_brain.model),
        "enabled": bool(setting.get("enabled", 1)),
        "monthly_budget_usd": monthly_budget,
        "input_cost_per_1k": float(setting.get("input_cost_per_1k", 0.0012) or 0.0012),
        "output_cost_per_1k": float(setting.get("output_cost_per_1k", 0.0024) or 0.0024),
        "base_url": db_base,
        "timeout_sec": timeout_out,
        "api_key_configured": key_cfg,
        "prompt_tokens": int(usage["prompt_tokens"] or 0),
        "completion_tokens": int(usage["completion_tokens"] or 0),
        "total_tokens": int(usage["total_tokens"] or 0),
        "estimated_cost_usd": round(estimated, 6),
        "budget_used_rate": rate,
    }


class BudgetSettingIn(BaseModel):
    model: str
    enabled: bool = True
    monthly_budget_usd: float = 50
    input_cost_per_1k: float = 0.0012
    output_cost_per_1k: float = 0.0024
    api_key: str | None = None
    base_url: str | None = None
    timeout_sec: float | None = None


@app.post("/api/admin/budget/model")
def admin_update_budget_setting(data: BudgetSettingIn) -> dict[str, Any]:
    dash = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM llm_settings ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()
    now = now_iso()
    if row:
        ex = dict(row)
        if data.api_key is not None and str(data.api_key).strip():
            api_key_final = str(data.api_key).strip()
        else:
            api_key_final = str(ex.get("api_key") or "").strip()
        base_final = (str(data.base_url or "").strip()) or (str(ex.get("base_url") or "").strip()) or dash
        if data.timeout_sec is not None:
            timeout_final = float(data.timeout_sec)
        else:
            try:
                timeout_final = float(ex.get("timeout_sec") or 8)
            except (TypeError, ValueError):
                timeout_final = 8.0
        cur.execute(
            """
            UPDATE llm_settings
            SET model=?, enabled=?, monthly_budget_usd=?, input_cost_per_1k=?, output_cost_per_1k=?,
                api_key=?, base_url=?, timeout_sec=?, updated_at=?
            WHERE id=?
            """,
            (
                data.model,
                1 if data.enabled else 0,
                data.monthly_budget_usd,
                data.input_cost_per_1k,
                data.output_cost_per_1k,
                api_key_final or None,
                base_final,
                timeout_final,
                now,
                ex["id"],
            ),
        )
    else:
        api_key_ins = str(data.api_key).strip() if data.api_key else ""
        base_ins = (str(data.base_url or "").strip()) or dash
        timeout_ins = float(data.timeout_sec) if data.timeout_sec is not None else 8.0
        cur.execute(
            """
            INSERT INTO llm_settings (
              provider, model, enabled, monthly_budget_usd, input_cost_per_1k, output_cost_per_1k,
              api_key, base_url, timeout_sec, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "bailian",
                data.model,
                1 if data.enabled else 0,
                data.monthly_budget_usd,
                data.input_cost_per_1k,
                data.output_cost_per_1k,
                api_key_ins or None,
                base_ins,
                timeout_ins,
                now,
            ),
        )
    conn.commit()
    conn.close()
    refresh_ai_brain_from_db()
    return {"message": "saved"}


@app.post("/api/ops/generate-strategies")
def ops_generate_strategies(
    batch_id: str,
    top_n: int = Query(default=300, ge=1, le=5000),
    sort_by: str = Query(default="qty"),
    candidate_source: str = Query(default="library"),
    use_ai: bool = Query(default=True),
    force_ai_only: bool = Query(default=False),
) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    order_sql = "total_qty DESC, total_gmv DESC"
    if sort_by == "gmv":
        order_sql = "total_gmv DESC, total_qty DESC"
    mains = cur.execute(
        f"""
        SELECT
          sku_id,
          MIN(product_name) AS product_name,
          MIN(category) AS category,
          AVG(price) AS price,
          AVG(cost) AS cost,
          SUM(qty) AS total_qty,
          SUM(gmv) AS total_gmv
        FROM uploaded_products
        WHERE batch_id=? AND role_hint='main'
        GROUP BY sku_id
        ORDER BY {order_sql}
        LIMIT ?
        """,
        (batch_id, top_n),
    ).fetchall()
    batch_candidates = cur.execute(
        """
        SELECT
          sku_id,
          MIN(product_name) AS product_name,
          MIN(category) AS category,
          AVG(price) AS price,
          AVG(cost) AS cost
        FROM uploaded_products
        WHERE batch_id=? AND role_hint='addon'
        GROUP BY sku_id
        """,
        (batch_id,),
    ).fetchall()
    library_candidates = cur.execute(
        "SELECT sku_id, product_name, category, original_price AS price, cost FROM products WHERE active=1 AND role='addon'"
    ).fetchall()
    if candidate_source == "batch":
        candidate_rows = batch_candidates
    elif candidate_source == "mixed":
        candidate_rows = batch_candidates + library_candidates
    else:
        candidate_rows = library_candidates
    if not candidate_rows:
        candidate_rows = batch_candidates or library_candidates

    strategy = get_latest_strategy()
    created = 0
    ai_count = 0
    rule_count = 0
    # Batch-level diversity: avoid all recommendations collapsing to the same SKU.
    selection_counts: dict[str, int] = {}
    skip_no_candidates = 0
    skip_exception = 0
    errors: list[dict[str, str]] = []
    now = now_iso()
    cur.execute("DELETE FROM bundle_recommendations WHERE batch_id=?", (batch_id,))
    for m in mains:
        policy = cur.execute(
            "SELECT * FROM policies WHERE category=? AND active=1 ORDER BY updated_at DESC LIMIT 1",
            (m["category"],),
        ).fetchone()
        if not policy:
            policy = {"logic_type": "慢病管理", "prompt_hint": "建议结合当前症状进行综合健康管理。", "margin_rate": 0.35}
        main_item = MainItem(
            sku_id=m["sku_id"],
            product_name=m["product_name"],
            category=m["category"],
            price=m["price"],
            cost=m["cost"],
        )
        candidate_pool = [
            CandidateItem(
                sku_id=a["sku_id"],
                product_name=a["product_name"],
                cost=a["cost"],
                original_price=a["price"],
                category=a["category"],
            )
            for a in candidate_rows
            if a["sku_id"] != m["sku_id"]
        ]
        if not candidate_pool:
            skip_no_candidates += 1
            continue
        try:
            result = _build_recommendation_result(
                main_item=main_item,
                user_id=None,
                candidates=candidate_pool,
                policy_like={
                    "logic_type": policy["logic_type"],
                    "prompt_hint": policy["prompt_hint"],
                    "margin_rate": policy["margin_rate"],
                },
                strategy=strategy,
                variant="A",
                prefer_ai=use_ai,
                force_ai_only=force_ai_only,
                selection_counts=selection_counts,
            )
        except ValueError as exc:
            if "候选池为空" in str(exc):
                skip_no_candidates += 1
            else:
                skip_exception += 1
                if len(errors) < 20:
                    errors.append({"sku_id": m["sku_id"], "product_name": m["product_name"], "reason": str(exc)})
            continue
        except Exception as exc:
            skip_exception += 1
            if len(errors) < 20:
                errors.append({"sku_id": m["sku_id"], "product_name": m["product_name"], "reason": str(exc) or "unexpected_error"})
            continue
        rec = result["recommendation"]
        src = rec.get("decision_trace", {}).get("source", "rule_engine")
        if src == "bailian_llm":
            ai_count += 1
        else:
            rule_count += 1
        # Update diversity counts after we know the final selected SKU.
        selection_counts[str(rec.get("selected_sku_id", ""))] = selection_counts.get(str(rec.get("selected_sku_id", "")), 0) + 1
        cur.execute(
            """
            INSERT INTO bundle_recommendations (
              batch_id, main_sku_id, main_product_name, main_category, selected_sku_id,
              selected_product_name, medical_logic, addon_price, projected_profit, sales_copy,
              decision_payload, status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, ?)
            """,
            (
                batch_id,
                m["sku_id"],
                m["product_name"],
                m["category"],
                rec["selected_sku_id"],
                rec["product_name"],
                rec["medical_logic"],
                rec["pricing_strategy"]["addon_price"],
                rec["projected_profit"],
                rec["sales_copy"],
                json.dumps(rec, ensure_ascii=False),
                now,
                now,
            ),
        )
        created += 1

    conn.commit()
    conn.close()
    return {
        "batch_id": batch_id,
        "generated_count": created,
        "top_n": top_n,
        "sort_by": sort_by,
        "candidate_source": candidate_source,
        "use_ai": use_ai,
        "ai_runtime_enabled": _is_ai_allowed(),
        "ai_generated_count": ai_count,
        "rule_generated_count": rule_count,
        "diagnostics": {
            "main_items_processed": len(mains),
            "candidate_pool_size": len(candidate_rows),
            "skip_no_candidates": skip_no_candidates,
            "skip_exception": skip_exception,
            "errors": errors,
        },
    }


@app.post("/api/ops/pricing/generate")
def ops_generate_pricing_assistant(
    batch_id: str,
    max_adjust_ratio: float = Query(default=0.10, ge=0.01, le=0.30),
    sales_weight: float = Query(default=0.6, ge=0.0, le=1.0),
    profit_weight: float = Query(default=0.4, ge=0.0, le=1.0),
    soft_margin_floor: float = Query(default=0.05, ge=0.0, le=0.5),
) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT
          product_name, manufacturer, sku_id, product_code, standard_code, category, business_mode, platform, merchant_name,
          price, cost, qty, gmv, revenue, gaap_profit, gaap_margin, order_cnt
        FROM uploaded_products
        WHERE batch_id=?
        """,
        (batch_id,),
    ).fetchall()
    if not rows:
        conn.close()
        raise HTTPException(status_code=404, detail="该批次暂无可用于定价的数据")

    # 归一权重，避免调用方传错导致偏移
    sw = max(float(sales_weight), 0.0)
    pw = max(float(profit_weight), 0.0)
    total_w = sw + pw
    if total_w <= 0:
        sw, pw = 0.6, 0.4
    else:
        sw, pw = sw / total_w, pw / total_w

    def _is_mp(mode: str) -> bool:
        m = (mode or "").upper()
        return "MP" in m

    grouped: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    for r in rows:
        d = dict(r)
        key = (
            str(d.get("product_name") or "").strip(),
            str(d.get("manufacturer") or "").strip(),
            str(d.get("product_code") or "").strip(),
            str(d.get("standard_code") or "").strip(),
            str(d.get("business_mode") or "").strip(),
            str(d.get("platform") or "").strip() + "|" + str(d.get("merchant_name") or "").strip(),
        )
        if key not in grouped:
            grouped[key] = {
                "product_name": key[0],
                "manufacturer": key[1],
                "product_code": key[2] or str(d.get("sku_id") or "").strip(),
                "standard_code": key[3],
                "category": str(d.get("category") or "").strip(),
                "business_mode": key[4],
                "platform": str(d.get("platform") or "").strip(),
                "merchant_name": str(d.get("merchant_name") or "").strip(),
                "gmv": 0.0,
                "revenue": 0.0,
                "gaap_profit": 0.0,
                "qty": 0.0,
                "order_cnt": 0.0,
                "cost_weighted_sum": 0.0,
            }
        g = grouped[key]
        qty = max(to_float(d.get("qty"), 0), 0.0)
        order_cnt = max(to_float(d.get("order_cnt"), 0), 0.0)
        gmv = to_float(d.get("gmv"), 0.0)
        revenue = to_float(d.get("revenue"), 0.0)
        gp = to_float(d.get("gaap_profit"), 0.0)
        cost = max(to_float(d.get("cost"), 0.0), 0.0)
        g["gmv"] += gmv
        g["revenue"] += revenue
        g["gaap_profit"] += gp
        g["qty"] += qty
        g["order_cnt"] += order_cnt
        g["cost_weighted_sum"] += cost * max(qty, 1.0)

    items: list[dict[str, Any]] = []
    for _, g in grouped.items():
        sales_base = max(g["order_cnt"], g["qty"], 1.0)
        current_price = g["gmv"] / sales_base if g["gmv"] else 0.0
        if current_price <= 0:
            continue
        avg_cost = g["cost_weighted_sum"] / max(g["qty"], 1.0)
        mp_mode = _is_mp(g["business_mode"])
        if mp_mode:
            profit_metric = (g["revenue"] / g["gmv"]) if abs(g["gmv"]) > 1e-9 else 0.0
            current_margin = (g["gaap_profit"] / g["revenue"]) if abs(g["revenue"]) > 1e-9 else 1.0
        else:
            if abs(g["revenue"]) > 1e-9:
                current_margin = g["gaap_profit"] / g["revenue"]
            else:
                current_margin = (current_price - avg_cost) / max(current_price, 1e-9)
            profit_metric = current_margin
        items.append(
            {
                **g,
                "sales_metric": sales_base,
                "current_price": current_price,
                "avg_cost": avg_cost,
                "profit_metric": profit_metric,
                "current_margin": current_margin,
                "mp_mode": mp_mode,
            }
        )

    if not items:
        conn.close()
        raise HTTPException(status_code=400, detail="该批次缺少可计算的 GMV/销量数据")

    peer_buckets_exact: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    peer_buckets_pid_mode: dict[tuple[str, str], list[dict[str, Any]]] = {}
    peer_buckets_cat_mode: dict[tuple[str, str], list[dict[str, Any]]] = {}
    peer_buckets_mode: dict[str, list[dict[str, Any]]] = {}
    for it in items:
        pid = (it["standard_code"] or it["product_code"] or it["product_name"]).strip()
        mode = it["business_mode"]
        category = (it.get("category") or "未分类").strip() or "未分类"
        bkey = (pid, it["manufacturer"], mode)
        peer_buckets_exact.setdefault(bkey, []).append(it)
        peer_buckets_pid_mode.setdefault((pid, mode), []).append(it)
        peer_buckets_cat_mode.setdefault((category, mode), []).append(it)
        peer_buckets_mode.setdefault(mode, []).append(it)

    now = now_iso()
    cur.execute("DELETE FROM pricing_recommendations WHERE batch_id=?", (batch_id,))
    created = 0
    up = 0
    down = 0
    keep = 0
    for it in items:
        pid = (it["standard_code"] or it["product_code"] or it["product_name"]).strip()
        mode = it["business_mode"]
        category = (it.get("category") or "未分类").strip() or "未分类"
        peers = peer_buckets_exact.get((pid, it["manufacturer"], mode), [])
        peer_source = "exact"
        if len(peers) < 3:
            peers = peer_buckets_pid_mode.get((pid, mode), peers)
            peer_source = "pid_mode"
        if len(peers) < 3:
            peers = peer_buckets_cat_mode.get((category, mode), peers)
            peer_source = "category_mode"
        if len(peers) < 3:
            peers = peer_buckets_mode.get(mode, peers) or peers
            peer_source = "mode_fallback"
        if not peers:
            peers = [it]
            peer_source = "self_only"
        peer_sales = sum(x["sales_metric"] for x in peers) / max(len(peers), 1)
        peer_profit = sum(x["profit_metric"] for x in peers) / max(len(peers), 1)
        # 四象限决策：
        # X轴：销量相对同品平均（好/差）
        # Y轴：利润相对同品平均（好/差）
        sales_delta = (it["sales_metric"] - peer_sales) / max(peer_sales, 1.0)
        profit_denom = max(abs(peer_profit), 0.01)
        profit_delta = (it["profit_metric"] - peer_profit) / profit_denom
        sales_delta = clamp(sales_delta, -1.0, 1.0)
        profit_delta = clamp(profit_delta, -1.0, 1.0)

        sales_good = sales_delta >= 0.0
        profit_good = profit_delta >= 0.0

        # 定义象限：
        # Q1：销量好 + 利润好（小幅提价）
        # Q2：销量好 + 利润差（提价修利润）
        # Q3：销量差 + 利润好（降价拉销量）
        # Q4：销量差 + 利润差（小幅降价先拉量，利润用保护线兜住）
        if sales_good and profit_good:
            quadrant = "Q1"
            delta = max_adjust_ratio * clamp(profit_delta, 0.0, 1.0) * profit_weight * 0.45
        elif sales_good and (not profit_good):
            quadrant = "Q2"
            delta = max_adjust_ratio * clamp(abs(profit_delta), 0.0, 1.0) * profit_weight * 0.75
        elif (not sales_good) and profit_good:
            quadrant = "Q3"
            delta = -max_adjust_ratio * clamp(abs(sales_delta), 0.0, 1.0) * sales_weight * 0.75
        else:
            quadrant = "Q4"
            delta = -max_adjust_ratio * (0.6 * clamp(abs(sales_delta), 0.0, 1.0) + 0.4 * clamp(abs(profit_delta), 0.0, 1.0)) * sales_weight * 0.65

        delta = clamp(delta, -max_adjust_ratio, max_adjust_ratio)
        # 避免“全体持平”：四象限已判断出方向时，给一个最小动作幅度（0.5%）
        min_move = min(0.005, max_adjust_ratio)
        if delta > 0:
            delta = max(delta, min_move)
        elif delta < 0:
            delta = min(delta, -min_move)
        # 只有极小变化才归零
        if abs(delta) < 0.003:
            delta = 0.0

        predicted_margin = it["current_margin"]
        if not it["mp_mode"]:
            if it["current_margin"] < soft_margin_floor and delta < 0:
                delta = max(delta, -0.02)
            suggested_try = max(0.01, it["current_price"] * (1 + delta))
            predicted_margin = (suggested_try - it["avg_cost"]) / max(suggested_try, 1e-9)
            if predicted_margin < soft_margin_floor and delta < 0:
                delta = delta * 0.5
                suggested_try = max(0.01, it["current_price"] * (1 + delta))
                predicted_margin = (suggested_try - it["avg_cost"]) / max(suggested_try, 1e-9)

        suggested_price = round(max(0.01, it["current_price"] * (1 + delta)), 2)
        if delta > 0.001:
            action = "建议小幅提价"
            up += 1
        elif delta < -0.001:
            action = "建议小幅降价"
            down += 1
        else:
            action = "建议维持现价"
            keep += 1

        reason = (
            f"{action}（销量权重{sw:.1f}/利润权重{pw:.1f}，幅度上限{max_adjust_ratio*100:.0f}%）；"
            f"四象限{quadrant}（销量{('好' if sales_good else '差')}、利润{('好' if profit_good else '差')}，"
            f"销量相对差{sales_delta:.3f}、利润相对差{profit_delta:.3f}，对标池{peer_source}:{len(peers)}）；"
            f"销量指标{it['sales_metric']:.0f}，利润指标{it['profit_metric']:.3f}。"
        )
        confidence = clamp(0.48 + 0.06 * min(len(peers), 5) + 0.25 * abs(delta), 0.5, 0.95)
        cur.execute(
            """
            INSERT INTO pricing_recommendations (
              batch_id, product_name, manufacturer, product_code, standard_code, business_mode, platform, merchant_name,
              current_price, suggested_price, delta_ratio, sales_metric, profit_metric, current_margin, predicted_margin,
              confidence, reason, status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'draft', ?, ?)
            """,
            (
                batch_id,
                it["product_name"],
                it["manufacturer"],
                it["product_code"],
                it["standard_code"],
                it["business_mode"],
                it["platform"],
                it["merchant_name"],
                round(it["current_price"], 2),
                suggested_price,
                round(delta, 4),
                round(it["sales_metric"], 2),
                round(it["profit_metric"], 4),
                round(it["current_margin"], 4),
                round(predicted_margin, 4),
                round(confidence, 3),
                reason,
                now,
                now,
            ),
        )
        created += 1

    conn.commit()
    conn.close()
    return {
        "batch_id": batch_id,
        "generated_count": created,
        "rule_config": {
            "sales_weight": round(sw, 3),
            "profit_weight": round(pw, 3),
            "max_adjust_ratio": max_adjust_ratio,
            "soft_margin_floor": soft_margin_floor,
        },
        "actions": {"up": up, "down": down, "keep": keep},
    }


@app.get("/api/ops/pricing/list")
def ops_list_pricing_recommendations(
    batch_id: str,
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    status: str | None = None,
) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    where = ["batch_id=?"]
    params: list[Any] = [batch_id]
    if status:
        where.append("status=?")
        params.append(status)
    where_sql = " AND ".join(where)
    total = cur.execute(
        f"SELECT COUNT(*) AS c FROM pricing_recommendations WHERE {where_sql}",
        tuple(params),
    ).fetchone()["c"]
    rows = cur.execute(
        f"""
        SELECT * FROM pricing_recommendations
        WHERE {where_sql}
        ORDER BY ABS(delta_ratio) DESC, id DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [limit, offset]),
    ).fetchall()
    conn.close()
    return {"total": total, "items": [dict(r) for r in rows]}


@app.post("/api/ops/pricing/{item_id}/confirm")
def ops_confirm_pricing(item_id: int) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("UPDATE pricing_recommendations SET status='confirmed', updated_at=? WHERE id=?", (now_iso(), item_id))
    conn.commit()
    conn.close()
    return {"message": "confirmed"}


@app.post("/api/ops/pricing/sync")
def ops_sync_pricing(batch_id: str) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT * FROM pricing_recommendations WHERE batch_id=? AND status='confirmed'",
        (batch_id,),
    ).fetchall()
    now = now_iso()
    synced = 0
    for r in rows:
        d = dict(r)
        key = "|".join(
            [
                str(d.get("standard_code") or ""),
                str(d.get("product_code") or ""),
                str(d.get("manufacturer") or ""),
                str(d.get("business_mode") or ""),
                str(d.get("platform") or ""),
                str(d.get("merchant_name") or ""),
            ]
        )
        if key == "|||||":
            key = f"name|{d.get('product_name','')}"
        cur.execute(
            """
            INSERT INTO pricing_rules (
              rule_key, product_name, manufacturer, product_code, standard_code, business_mode, platform, merchant_name,
              current_price, suggested_price, delta_ratio, current_margin, predicted_margin, reason, active, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(rule_key) DO UPDATE SET
              product_name=excluded.product_name,
              manufacturer=excluded.manufacturer,
              product_code=excluded.product_code,
              standard_code=excluded.standard_code,
              business_mode=excluded.business_mode,
              platform=excluded.platform,
              merchant_name=excluded.merchant_name,
              current_price=excluded.current_price,
              suggested_price=excluded.suggested_price,
              delta_ratio=excluded.delta_ratio,
              current_margin=excluded.current_margin,
              predicted_margin=excluded.predicted_margin,
              reason=excluded.reason,
              active=1,
              updated_at=excluded.updated_at
            """,
            (
                key,
                d.get("product_name"),
                d.get("manufacturer"),
                d.get("product_code"),
                d.get("standard_code"),
                d.get("business_mode"),
                d.get("platform"),
                d.get("merchant_name"),
                d.get("current_price"),
                d.get("suggested_price"),
                d.get("delta_ratio"),
                d.get("current_margin"),
                d.get("predicted_margin"),
                d.get("reason"),
                now,
            ),
        )
        cur.execute("UPDATE pricing_recommendations SET status='published', updated_at=? WHERE id=?", (now, d["id"]))
        synced += 1
    conn.commit()
    conn.close()
    return {"batch_id": batch_id, "synced_count": synced}


@app.get("/api/ops/pricing/export")
def ops_export_pricing_csv(batch_id: str, status: str | None = None) -> Response:
    conn = db_conn()
    cur = conn.cursor()
    where = ["batch_id=?"]
    params: list[Any] = [batch_id]
    if status:
        where.append("status=?")
        params.append(status)
    where_sql = " AND ".join(where)
    rows = cur.execute(
        f"""
        SELECT
          id, product_name, manufacturer, product_code, standard_code, business_mode, platform, merchant_name,
          current_price, suggested_price, delta_ratio, current_margin, predicted_margin, reason, status
        FROM pricing_recommendations
        WHERE {where_sql}
        ORDER BY ABS(delta_ratio) DESC, id DESC
        """,
        tuple(params),
    ).fetchall()
    conn.close()
    header = [
        "ID",
        "产品名称",
        "厂家",
        "商品编码",
        "标品主码",
        "业务模式",
        "平台",
        "商家",
        "当前单价",
        "建议单价",
        "调价幅度",
        "当前毛利率",
        "预测毛利率",
        "建议说明",
        "状态",
    ]
    lines = [",".join(header)]
    for r in rows:
        d = dict(r)
        vals = [
            d.get("id"),
            d.get("product_name", ""),
            d.get("manufacturer", ""),
            d.get("product_code", ""),
            d.get("standard_code", ""),
            d.get("business_mode", ""),
            d.get("platform", ""),
            d.get("merchant_name", ""),
            f"{to_float(d.get('current_price'), 0):.2f}",
            f"{to_float(d.get('suggested_price'), 0):.2f}",
            f"{to_float(d.get('delta_ratio'), 0)*100:.2f}%",
            f"{to_float(d.get('current_margin'), 0)*100:.2f}%",
            f"{to_float(d.get('predicted_margin'), 0)*100:.2f}%",
            d.get("reason", ""),
            d.get("status", ""),
        ]
        esc = [f"\"{str(x).replace('\"', '\"\"')}\"" for x in vals]
        lines.append(",".join(esc))
    csv_text = "\ufeff" + "\n".join(lines)
    filename = f"pricing_recommendations_{batch_id}.csv"
    return Response(
        content=csv_text,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/ops/strategies")
def ops_list_strategies(
    batch_id: str,
    status: str | None = None,
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    where = ["batch_id=?"]
    params: list[Any] = [batch_id]
    if status:
        where.append("status=?")
        params.append(status)
    where_sql = " AND ".join(where)
    total = cur.execute(
        f"SELECT COUNT(*) AS c FROM bundle_recommendations WHERE {where_sql}",
        tuple(params),
    ).fetchone()["c"]
    rows = cur.execute(
        f"SELECT * FROM bundle_recommendations WHERE {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?",
        tuple(params + [limit, offset]),
    ).fetchall()
    conn.close()
    items = []
    for r in rows:
        d = dict(r)
        source = "rule_engine"
        try:
            payload = json.loads(d.get("decision_payload") or "{}")
            source = payload.get("decision_trace", {}).get("source", "rule_engine")
        except Exception:
            source = "rule_engine"
        d["source"] = source
        d["source_label"] = "百炼AI" if source == "bailian_llm" else "规则引擎"
        items.append(d)
    return {"total": total, "items": items}


@app.post("/api/ops/strategies/{item_id}/confirm")
def ops_confirm_strategy(item_id: int) -> dict[str, str]:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute("UPDATE bundle_recommendations SET status='confirmed', updated_at=? WHERE id=?", (now_iso(), item_id))
    conn.commit()
    conn.close()
    return {"message": "confirmed"}


@app.post("/api/ops/sync")
def ops_sync_confirmed(batch_id: str) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT * FROM bundle_recommendations WHERE batch_id=? AND status='confirmed'",
        (batch_id,),
    ).fetchall()
    now = now_iso()
    synced = 0
    for r in rows:
        cur.execute(
            """
            INSERT INTO bundle_rules (
              main_sku_id, main_product_name, selected_sku_id, selected_product_name,
              addon_price, medical_logic, sales_copy, active, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(main_sku_id) DO UPDATE SET
              main_product_name=excluded.main_product_name,
              selected_sku_id=excluded.selected_sku_id,
              selected_product_name=excluded.selected_product_name,
              addon_price=excluded.addon_price,
              medical_logic=excluded.medical_logic,
              sales_copy=excluded.sales_copy,
              active=1,
              updated_at=excluded.updated_at
            """,
            (
                r["main_sku_id"],
                r["main_product_name"],
                r["selected_sku_id"],
                r["selected_product_name"],
                r["addon_price"],
                r["medical_logic"],
                r["sales_copy"],
                now,
            ),
        )
        cur.execute("UPDATE bundle_recommendations SET status='published', updated_at=? WHERE id=?", (now, r["id"]))
        synced += 1
    conn.commit()
    conn.close()
    return {"batch_id": batch_id, "synced_count": synced}


@app.get("/api/ops/workbench")
def ops_workbench(batch_id: str) -> dict[str, Any]:
    conn = db_conn()
    cur = conn.cursor()
    total = cur.execute("SELECT COUNT(*) AS c FROM bundle_recommendations WHERE batch_id=?", (batch_id,)).fetchone()["c"]
    draft = cur.execute(
        "SELECT COUNT(*) AS c FROM bundle_recommendations WHERE batch_id=? AND status='draft'",
        (batch_id,),
    ).fetchone()["c"]
    confirmed = cur.execute(
        "SELECT COUNT(*) AS c FROM bundle_recommendations WHERE batch_id=? AND status='confirmed'",
        (batch_id,),
    ).fetchone()["c"]
    published = cur.execute(
        "SELECT COUNT(*) AS c FROM bundle_recommendations WHERE batch_id=? AND status='published'",
        (batch_id,),
    ).fetchone()["c"]
    sku_count = cur.execute(
        "SELECT COUNT(DISTINCT sku_id) AS c FROM uploaded_products WHERE batch_id=? AND role_hint='main'",
        (batch_id,),
    ).fetchone()["c"]
    top_qty_rows = cur.execute(
        """
        SELECT sku_id, MIN(product_name) AS product_name, SUM(qty) AS total_qty, SUM(gmv) AS total_gmv
        FROM uploaded_products
        WHERE batch_id=? AND role_hint='main'
        GROUP BY sku_id
        ORDER BY total_qty DESC, total_gmv DESC
        LIMIT 10
        """,
        (batch_id,),
    ).fetchall()
    pricing_total = cur.execute("SELECT COUNT(*) AS c FROM pricing_recommendations WHERE batch_id=?", (batch_id,)).fetchone()["c"]
    pricing_up = cur.execute(
        "SELECT COUNT(*) AS c FROM pricing_recommendations WHERE batch_id=? AND delta_ratio>0.001",
        (batch_id,),
    ).fetchone()["c"]
    pricing_down = cur.execute(
        "SELECT COUNT(*) AS c FROM pricing_recommendations WHERE batch_id=? AND delta_ratio<-0.001",
        (batch_id,),
    ).fetchone()["c"]
    pricing_confirmed = cur.execute(
        "SELECT COUNT(*) AS c FROM pricing_recommendations WHERE batch_id=? AND status='confirmed'",
        (batch_id,),
    ).fetchone()["c"]
    pricing_published = cur.execute(
        "SELECT COUNT(*) AS c FROM pricing_recommendations WHERE batch_id=? AND status='published'",
        (batch_id,),
    ).fetchone()["c"]
    conn.close()
    return {
        "batch_id": batch_id,
        "sku_count": sku_count,
        "total": total,
        "draft": draft,
        "confirmed": confirmed,
        "published": published,
        "pricing_total": pricing_total,
        "pricing_up": pricing_up,
        "pricing_down": pricing_down,
        "pricing_confirmed": pricing_confirmed,
        "pricing_published": pricing_published,
        "top_products": [dict(r) for r in top_qty_rows],
    }
